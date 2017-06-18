#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Mazama_BP_2017.py

A script to convert the British Petroleum Statistical Review Excel workbook into
a series of standardized ASCII CSV files for ingest by other software.

"""

import sys
reload(sys)  # Reload does the trick!
sys.setdefaultencoding('UTF8')
from openpyxl import load_workbook
from Mazama_Countries import CountryTranslator
import pprint
import collections

RELEASE_YEAR = 2017
CSV_START_YEAR = 1965

# Returns a data dictionary filled with the contents of a BP Stat Review data sheet.
def get_data_dictionary(sheet,rowrange,colrange):
    result = {}

    ct = CountryTranslator('BP_2015')
    import ignore_me as ign_me
    ignore_me = ign_me.ignore_me

    for r in rowrange:
        BP_country_code = sheet.cell(row=r,column=1+0).value
        if r == 3:
            BP_country_code = 'YEAR'
        
        if BP_country_code == None:
            continue
        # NOTE:  Remove pound sterling sign (u'\xa3') from 'Non-OPEC'
        # NOTE:  Strip initial and trailing space
        BP_country_code = BP_country_code.replace(u'\xa3','').strip()
        if BP_country_code in ignore_me:
            continue
        try:
            MZM_code = ct.get_MZM_code(BP_country_code)
            ###print("DEBUG: MZM_code = '%s'" % (MZM_code))
            # print "%s => %s" % (BP_country_code, MZM_code) 
            result[MZM_code] = []
            
            for c in colrange:
                value = sheet.cell(row=r,column=c).value
                data_type = sheet.cell(row=r,column=c).data_type
                
                if data_type == 'n':
                    if r == 3:
                        result["YEAR"].append(value)
                    else:
                        result[MZM_code].append(value)
                        
                elif data_type == 's':
                    if value == u'-':
                        result[MZM_code].append(0.0)
                    elif value == u'^':
                        result[MZM_code].append(0.0)
                    elif value == u'n/a':
                        result[MZM_code].append("na")
                    else:
                        try:
                            value = float(value)
                            result[MZM_code].append(value)
                        except Exception, e:
                            print("ERROR: " + str(e) + ": cell value \"" + value + "\" is not handled.")
                            sys.exit(1)
                    
                else:
                    print("UNKNOWN data_type %d" % (data_type))
                    sys.exit(1)
            
        except Exception, e:
            ###error_text = "ERROR in get_data_dictionary:  BP_country_code = \n\t%s\nerror = \n\t%s" % (BP_country_code,e)
            ###error_text = "ERROR: %s\nu\"%s\"," % (e, BP_country_code)
            error_text = "                 u\"%s\"," % (BP_country_code)
            print(error_text)
            ###exit(1)

            pass

    return(result)


########################################
# write_data_as_csv
#
# The csv file is organized as Year (row) X MZM_code (col)

def write_data_as_csv(filename, Data, rowrange, rounding, data_start_year=1965):

    # Get a sorted list of keys
    MZM_codes = Data.keys()
    MZM_codes.sort()
    MZM_codes.remove("YEAR")

    # Write out the header line
    filename.write("\"YEAR\"")
    for MZM_code in MZM_codes:
        filename.write(",\"" + MZM_code + "\"")
    filename.write("\n")

    # Prefill with 'na' as needed
    for row in range(CSV_START_YEAR, data_start_year):
        filename.write(str(row))
        for MZM_code in MZM_codes:
            filename.write(",\"na\"")
        filename.write("\n")

    # For each row, continue by writing out the year and all values
    for row in rowrange:
        filename.write(str(Data["YEAR"][row]))
        for MZM_code in MZM_codes:
            try:
                filename.write("," + str(round(Data[MZM_code][row],rounding)))
            except TypeError:
                filename.write(",\"na\"")
        filename.write("\n")


##############################################################################
# Main program
#
def main():

    stat_review = 'BP_2017.xlsx'

    print("Loading %s ..." % (stat_review))

    try:
        workbook = load_workbook(filename=stat_review)
    except:
        print("*** Open failed: %s: %s" % (sys.exc_info()[:2]))

    print("Successfully opened workbook.")

    title_to_filename = {
        #'Contents',
        #'Primary Energy Consumption',
        #'Primary Energy - Cons by fuel',
        #'Oil - Proved reserves',
        'Oil - Proved reserves history' : 'BP_2017_oil_history_gb',
        'Oil Production - Barrels' : "BP_2017_oil_production_bbl",
        'Oil Production - Tonnes' : "BP_2017_oil_production_mtoe",
        'Oil Consumption -  Barrels': "BP_2017_oil_consumption_bbl",
        'Oil Consumption - Tonnes': "BP_2017_oil_consumption_mtoe",
        #'Oil - Regional Consumption ',
        #'Oil - Spot crude prices',
        #'Oil - Crude prices since 1861',
        #'Oil - Refinery throughput',
        #'Oil - Refinery capacities',
        #'Oil - Regional refining margins',
        #'Oil - Trade movements',
        #'Oil - Inter-area movements ',
        #'Oil - Trade 2015 - 2016',
        #'Gas - Proved reserves',
        'Gas - Proved reserves history ' : 'BP_2017_gas_history_trillion_cubic_metres',
        'Gas Production - Bcm' : "BP_2017_gas_production_m3",
        'Gas Production - Bcf' : "BP_2017_gas_production_ft3",
        'Gas Production - Mtoe' : "BP_2017_gas_production_mtoe",
        'Gas Consumption - Bcm' : "BP_2017_gas_consumption_m3",
        'Gas Consumption - Bcf' : "BP_2017_gas_consumption_ft3",
        'Gas Consumption - Mtoe' : "BP_2017_gas_consumption_mtoe",
        # 'Gas - Trade - pipeline',
        # 'Gas - Trade movements LNG',
        # 'Gas - Trade 2015-2016',
        # 'Gas - Prices ',
        # 'Coal - Reserves',
        # 'Coal - Prices',
        'Coal Production - Tonnes' : "BP_2017_coal_production_ton",
        'Coal Production - Mtoe' : "BP_2017_coal_production_mtoe",
        'Coal Consumption -  Mtoe' : "BP_2017_coal_consumption_mtoe",
        'Nuclear Consumption - TWh' : "BP_2017_nuclear_consumption_twh",
        'Nuclear Consumption - Mtoe' : "BP_2017_nuclear_consumption_mtoe",
        'Hydro Consumption - TWh' : "BP_2017_hydro_consumption_twh",
        'Hydro Consumption - Mtoe': "BP_2017_hydro_consumption_mtoe",
        'Other renewables -TWh' : "BP_2017_renewables_consumption_twh",
        'Other renewables - Mtoe' : "BP_2017_renewables_consumption_mtoe",
        'Solar Consumption - TWh' : "BP_2017_solar_consumption_twh",
        'Solar Consumption - Mtoe': "BP_2017_solar_consumption_mtoe",
        'Wind Consumption - TWh ': "BP_2017_wind_consumption_twh",
        'Wind Consumption - Mtoe': "BP_2017_wind_consumption_mtoe",
        #'Geo Biomass Other - TWh' : "BP_2017_geo_biomass_other_twh",
        # 'Geo Biomass Other - Mtoe',
        # 'Biofuels Production - Kboed',
        # 'Biofuels Production - Ktoe',
        # 'Electricity Generation ',
        'Carbon Dioxide Emissions' : 'BP_2017_co2_emissions',
        # 'Geothermal capacity',
        # 'Solar capacity',
        # 'Wind capacity',
        # 'Approximate conversion factors',
        # 'Definitions'
    }

    print "Verifying expected worksheets is present ..."
    title_to_worksheet = dict([(sheet.title, sheet) for sheet in workbook.worksheets])
    for title in title_to_filename.keys():
        if title in title_to_worksheet:
            #print "%s => OK" % (title)
            pass
        else:
            pprint.pprint([sheet.title for sheet in workbook.worksheets])
            print "missing worksheet %s" % title
            sys.exit(1)

    import collections

    #sheet_indices = [sheet_indices[8]]
    for title in title_to_filename.keys():
        file_name = title_to_filename[title] + ".csv"
        sheet = title_to_worksheet[title]
        sheet_title = sheet.cell(row=1+0,column=1+0).value.replace('*','').rstrip()
        title = sheet.cell(row=1+0,column=1+0).value.replace('*','').rstrip()
        #print "title => %s => %s" % (title, sheet_title)
        units = sheet.cell(row=1+2,column=1+0).value.lower()

        # Determine rows and columns to read
        data_start_year = sheet.cell(row=1+2,column=1+1).value
        #print "data_start_year = %s" % (str(data_start_year))
        col_hi = RELEASE_YEAR - data_start_year + 1
        colrange = range(1+1,col_hi+1)
        rowrange = range(1,100) # Rowrange is larger than needed and rows that don't have country names will be skipped

        #print ("Converting %s (%s)" % (title,units)).ljust(85),
        # file_name = file_name.replace("_2017_", "_")
        # file_name = file_name.replace("_renewables_", "_other_renewables_")
        #print "=> %s ..." % (file_name)
        file = open("./" + file_name,'w')
        file.write("title         = ASCII CSV version of worksheet \"%s\" from the 2017 British Petroleum Statistical Review\n" % (title))
        file.write("file URL      = http://mazamascience.com/Data/Energy/BP/2017/%s\n" % (file_name))
        file.write("original data = http://www.bp.com/content/dam/bp/en/corporate/excel/energy-economics/statistical-review-2017/bp-statistical-review-of-world-energy-2017-underpinning-data.xlsx\n")
        file.write("country codes = ISO3166-1 two-letter codes or 'BP_~~~' for non-standard BP groupings (e.g. BP_TNA = Total North America)\n")
        file.write("units         = %s\n" % (units))
        file.write("\n")
        Data = get_data_dictionary(sheet,rowrange,colrange)
        rowrange = range(0,col_hi-1)
        rounding = 3
        write_data_as_csv(file,Data,rowrange,rounding,data_start_year)
        file.close()

        file = open('./' + str(RELEASE_YEAR) + '/' + file_name, 'w')
        write_data_as_csv(file,Data,rowrange,rounding,data_start_year)
        file.close()

################################################################################

if __name__ == "__main__":
    main()
