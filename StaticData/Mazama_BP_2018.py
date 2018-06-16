#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Mazama_BP_2017.py

A script to convert the British Petroleum Statistical Review Excel workbook into
a series of standardized ASCII CSV files for ingest by other software.

"""

import sys
reload(sys)  # Reload does the trick!
sys.setdefaultencoding('UTF8')
from openpyxl import load_workbook
from Mazama_Countries import CountryTranslator
import pprint
import collections

RELEASE_YEAR = 2018
CSV_START_YEAR = 1965

seen_errors = []

# Returns a data dictionary filled with the contents of a BP Stat Review data sheet.
def get_data_dictionary(sheet,rowrange,colrange):
    result = {}
    global seen_errors

    ct = CountryTranslator('BP_2015')
    import ignore_me as ign_me
    ignore_me = ign_me.ignore_me

    for r in rowrange:
        BP_country_code = sheet.cell(row=r,column=1+0).value
        if r == 3:
            BP_country_code = 'YEAR'
        
        if BP_country_code == None:
            continue
        # NOTE:  Remove pound sterling sign (u'\xa3') from 'Non-OPEC'
        # NOTE:  Strip initial and trailing space
        BP_country_code = BP_country_code.replace(u'\xa3','').strip()
        if BP_country_code in ignore_me:
            continue
        try:
            MZM_code = ct.get_MZM_code(BP_country_code)
            ###print("DEBUG: MZM_code = '%s'" % (MZM_code))
            # print "%s => %s" % (BP_country_code, MZM_code) 
            result[MZM_code] = []
            
            for c in colrange:
                value = sheet.cell(row=r,column=c).value
                data_type = sheet.cell(row=r,column=c).data_type
                
                if data_type == 'n':
                    if r == 3:
                        result["YEAR"].append(value)
                    else:
                        result[MZM_code].append(value)
                        
                elif data_type == 's':
                    if value == u'-':
                        result[MZM_code].append(0.0)
                    elif value == u'^':
                        result[MZM_code].append(0.0)
                    elif value == u'n/a':
                        result[MZM_code].append("na")
                    else:
                        try:
                            value = float(value)
                            result[MZM_code].append(value)
                        except Exception, e:
                            print("ERROR: " + str(e) + ": cell value \"" + value + "\" is not handled.")
                            sys.exit(1)
                    
                else:
                    print("UNKNOWN data_type %d" % (data_type))
                    sys.exit(1)
            
        except Exception, e:
            ###error_text = "ERROR in get_data_dictionary:  BP_country_code = \n\t%s\nerror = \n\t%s" % (BP_country_code,e)
            ###error_text = "ERROR: %s\nu\"%s\"," % (e, BP_country_code)
            error_text = "                 u\"%s\"," % (BP_country_code)
            if not BP_country_code in seen_errors:
                seen_errors.append(BP_country_code)
                print(error_text)
            ###exit(1)

            pass

    return(result)


########################################
# write_data_as_csv
#
# The csv file is organized as Year (row) X MZM_code (col)

def write_data_as_csv(filename, Data, rowrange, rounding, data_start_year=1965):

    # Get a sorted list of keys
    MZM_codes = Data.keys()
    MZM_codes.sort()
    MZM_codes.remove("YEAR")

    # Write out the header line
    filename.write("\"YEAR\"")
    for MZM_code in MZM_codes:
        filename.write(",\"" + MZM_code + "\"")
    filename.write("\n")

    # Prefill with 'na' as needed
    for row in range(CSV_START_YEAR, data_start_year):
        filename.write(str(row))
        for MZM_code in MZM_codes:
            filename.write(",\"na\"")
        filename.write("\n")

    # For each row, continue by writing out the year and all values
    for row in rowrange:
        filename.write(str(Data["YEAR"][row]))
        for MZM_code in MZM_codes:
            try:
                filename.write("," + str(round(Data[MZM_code][row],rounding)))
            except TypeError:
                filename.write(",\"na\"")
        filename.write("\n")


##############################################################################
# Main program
#
def main():

    stat_review = 'BP_' + str(RELEASE_YEAR) + '.xlsx'

    print("Loading %s ..." % (stat_review))

    try:
        workbook = load_workbook(filename=stat_review)
    except:
        print("*** Open failed: %s: %s" % (sys.exc_info()[:2]))
        sys.exit(1)

    print("Successfully opened workbook.")

    def fname(postfix):
        return "BP_" + str(RELEASE_YEAR) + "_" + postfix

    import title_to_filename as t2f
    title_to_filename = dict([(k, fname(v)) for (k,v) in t2f.title_to_filename.items()])

    print "Verifying expected worksheets is present ..."
    title_to_worksheet = dict([(sheet.title, sheet) for sheet in workbook.worksheets])
    for title in title_to_filename.keys():
        if title in title_to_worksheet:
            #print "%s => OK" % (title)
            pass
        else:
            elems = [sheet.title for sheet in workbook.worksheets]
            elems.sort()
            pprint.pprint(elems)
            print "missing worksheet >%s<" % title
            for candidate in elems:
                found = True
                for part in title.split(" "):
                    if part not in candidate:
                        found = False
                if found:
                    print "Did you mean >%s< ?" % (candidate)
            sys.exit(1)

    import collections

    #sheet_indices = [sheet_indices[8]]
    for title in title_to_filename.keys():
        #print "doing %s" % (title)
        file_name = title_to_filename[title] + ".csv"
        sheet = title_to_worksheet[title]
        sheet_title = sheet.cell(row=1+0,column=1+0).value.replace('*','').rstrip()
        title = sheet.cell(row=1+0,column=1+0).value.replace('*','').rstrip()
        #print "title => %s => %s" % (title, sheet_title)
        units = sheet.cell(row=1+2,column=1+0).value.lower()

        # Determine rows and columns to read
        data_start_year = sheet.cell(row=1+2,column=1+1).value
        #print "data_start_year = %s" % (str(data_start_year))
        col_hi = RELEASE_YEAR - data_start_year + 1
        colrange = range(1+1,col_hi+1)
        rowrange = range(1,100) # Rowrange is larger than needed and rows that don't have country names will be skipped

        #print ("Converting %s (%s)" % (title,units)).ljust(85),
        # file_name = file_name.replace("_' + str(RELEASE_YEAR) + '_", "_")
        # file_name = file_name.replace("_renewables_", "_other_renewables_")
        #print "=> %s ..." % (file_name)
        file = open("./" + file_name,'w')
        file.write(("title         = ASCII CSV version of worksheet \"%s\" from the " + str(RELEASE_YEAR) + " British Petroleum Statistical Review\n") % (title))
        file.write(("file URL      = http://mazamascience.com/Data/Energy/BP/" + str(RELEASE_YEAR) + "/%s\n") % (file_name))
        file.write("original data = http://www.bp.com/content/dam/bp/en/corporate/excel/energy-economics/statistical-review-" + str(RELEASE_YEAR) + "/bp-statistical-review-of-world-energy-" + str(RELEASE_YEAR) + "-underpinning-data.xlsx\n")
        file.write("country codes = ISO3166-1 two-letter codes or 'BP_~~~' for non-standard BP groupings (e.g. BP_TNA = Total North America)\n")
        file.write("units         = %s\n" % (units))
        file.write("\n")
        BP_DIAGRAMS_FILES = ['co2_emissions',
                             'coal_consumption_mtoe',
                             'coal_production_mtoe',
                             'gas_consumption_mtoe',
                             'gas_production_mtoe',
                             'hydro_consumption_mtoe',
                             'nuclear_consumption_mtoe',
                             'oil_consumption_mtoe',
                             'oil_production_mtoe',
                             'renewables_consumption_mtoe',
                             'solar_consumption_mtoe',
                             'wind_consumption_mtoe']
        #BP_DIAGRAMS_TRANSLATE = {'renewables_consumption_mtoe' }
        Data = get_data_dictionary(sheet,rowrange,colrange)
        rowrange = range(0,col_hi-1)
        rounding = 3
        write_data_as_csv(file,Data,rowrange,rounding,data_start_year)
        file.close()

        file = open('./' + str(RELEASE_YEAR) + '/' + file_name, 'w')
        write_data_as_csv(file,Data,rowrange,rounding,data_start_year)
        file.close()
        for candidate_file in BP_DIAGRAMS_FILES:
          if candidate_file in file_name:
            dest_file = "../../bp-diagrams/data/" + file_name.replace(("_" + str (RELEASE_YEAR) + ""), '').replace("BP_renewables_consumption_mtoe", "BP_other_renewables_consumption_mtoe").replace("BP_co2_emissions_mt", "BP_co2_emissions")
            print "copying file %s to %s" % (file_name, dest_file)
            import shutil
            shutil.copy(file_name, dest_file)

################################################################################

if __name__ == "__main__":
    main()
